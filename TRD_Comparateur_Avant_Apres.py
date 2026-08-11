#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
TRD / IEA - Comparateur Avant / Après pour consolidations Radio.

Le script demande :
  1) le fichier AVANT
  2) le fichier APRES
  3) où enregistrer le résultat

Résultat :
  - Comparatif_TRD : détail secteur + bandes + TRD/IEA
  - Dashboard_TRD  : synthèse des résultats
  - Parametres_TRD : seuils utilisés

Installation :
    pip install openpyxl

Exécution :
    python TRD_Comparateur_Avant_Apres.py

Mode ligne de commande :
    python TRD_Comparateur_Avant_Apres.py ^
        --before "Consolidation_0708.xlsx" ^
        --after "Consolidation_1008.xlsx" ^
        --output "Resultat_TRD.xlsx"
"""

import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\nERREUR : openpyxl n'est pas installé.")
    print("Exécuter : pip install openpyxl\n")
    sys.exit(1)


# ============================================================
# PARAMETRES
# ============================================================

DL_THRESHOLDS = {
    "L1800": 10.0,   # Mbps
    "L2100": 5.0,
    "L2600": 10.0,
    "L800": 5.0,
}

PRB_THRESHOLD = 70.0       # %
WEIGHT_TRD_DL = 0.60
WEIGHT_TRD_PRB = 0.40

IEA_VERY_EFFECTIVE = 0.80
IEA_EFFECTIVE = 0.50
IEA_PARTIAL = 0.20
IEA_REGRESSION = -0.20

BANDS = ("L1800", "L2100", "L2600", "L800")

# Arrêt de lecture après N lignes consécutives sans secteur.
# Utile pour les fichiers avec 1 048 576 lignes formatées mais vides.
MAX_CONSECUTIVE_EMPTY_ROWS = 200


# ============================================================
# UTILITAIRES
# ============================================================

def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(value):
    s = clean_text(value).lower().replace("_", " ")
    return re.sub(r"\s+", " ", s)


def normalize_sector(value):
    return re.sub(r"\s+", " ", clean_text(value))


def to_float(value):
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").strip())
    except Exception:
        return None


def safe_get(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def band_set(value):
    text = clean_text(value).upper()
    result = set()
    for band in BANDS:
        digits = band[1:]
        if re.search(rf"(?<!\d){digits}(?!\d)", text):
            result.add(band)
    return result


def safe_gain(before, after):
    b = to_float(before)
    a = to_float(after)
    if b is None or a is None or b == 0:
        return None
    return a / b - 1.0


def prb_relief(before, after):
    b = to_float(before)
    a = to_float(after)
    if b is None or a is None:
        return None
    return b - a


def classify_iea(iea):
    if iea is None:
        return "Non mesurable"
    if iea >= IEA_VERY_EFFECTIVE:
        return "Très efficace"
    if iea >= IEA_EFFECTIVE:
        return "Efficace"
    if iea >= IEA_PARTIAL:
        return "Amélioration partielle"
    if iea < IEA_REGRESSION:
        return "Régression"
    return "Impact faible / non significatif"


def build_conclusion(degradation_status, action_status, iea):
    q = clean_text(degradation_status)

    if q.lower() == "normalisé":
        return f"Normalisé – {action_status}"

    if q.lower() == "persiste":
        if iea is None:
            return "Persiste – impact non mesurable"
        if iea >= IEA_EFFECTIVE:
            return "Persiste malgré amélioration forte"
        if iea >= IEA_PARTIAL:
            return "Persiste – amélioration partielle"
        if iea < IEA_REGRESSION:
            return "Persiste – régression"
        return "Persiste – impact faible"

    return f"{q} – {action_status}" if q else action_status


# ============================================================
# LECTURE OPTIMISEE DES FICHIERS
# ============================================================

def find_analysis_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "analyse":
            return wb[name]

    # Recherche limitée aux premières lignes.
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            if "secteurs" in [normalize_text(v) for v in row]:
                return ws

    raise ValueError(
        "Impossible de trouver l'onglet 'Analyse' ou une feuille contenant 'Secteurs'."
    )


def read_top_rows(ws, n=20):
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=n, values_only=True):
        rows.append(tuple(row))
    return rows


def find_header_row(top_rows):
    for r_idx, row in enumerate(top_rows):
        vals = [normalize_text(v) for v in row]
        if "secteurs" in vals and (
            "responsabilité" in vals
            or "qualif" in vals
            or "action déploiement" in vals
        ):
            return r_idx  # 0-based dans top_rows

    raise ValueError("Impossible d'identifier la ligne d'en-tête.")


def find_column(header_row, names, required=True):
    wanted = [normalize_text(x) for x in names]
    headers = [normalize_text(v) for v in header_row]

    for idx, h in enumerate(headers):
        if h in wanted:
            return idx

    for idx, h in enumerate(headers):
        for w in wanted:
            if w and h and (w in h or h in w):
                return idx

    if required:
        raise ValueError("Colonne introuvable : " + " / ".join(names))
    return None


def find_band_groups(top_rows, header_idx):
    """
    Cherche L1800/L2100/L2600/L800 au-dessus de la ligne d'en-tête,
    puis localise Débit 4G DL et Usage_PRB_DL_BH dans chaque groupe.
    """
    starts = {}

    for r in range(header_idx):
        row = top_rows[r]
        for c, value in enumerate(row):
            v = clean_text(value).upper().replace(" ", "")
            if v in BANDS:
                starts[v] = c

    if not starts:
        raise ValueError(
            "Impossible d'identifier les groupes L1800/L2100/L2600/L800."
        )

    ordered = sorted(starts.items(), key=lambda x: x[1])
    header = top_rows[header_idx]
    groups = {}

    for i, (band, start_col) in enumerate(ordered):
        end_col = (
            ordered[i + 1][1] - 1
            if i + 1 < len(ordered)
            else min(start_col + 5, len(header) - 1)
        )

        dl_col = None
        prb_col = None

        for c in range(start_col, end_col + 1):
            h = normalize_text(header[c] if c < len(header) else None)

            if "débit" in h and "dl" in h:
                dl_col = c

            if "prb" in h:
                prb_col = c

        groups[band] = {
            "dl": dl_col,
            "prb": prb_col,
        }

    missing = [
        band for band in BANDS
        if band not in groups
        or groups[band]["dl"] is None
        or groups[band]["prb"] is None
    ]

    if missing:
        raise ValueError(
            "Débit/PRB introuvable pour : " + ", ".join(missing)
        )

    return groups


def detect_structure(top_rows):
    header_idx = find_header_row(top_rows)
    header = top_rows[header_idx]

    structure = {
        "header_idx": header_idx,
        "header_excel_row": header_idx + 1,
        "sector": find_column(header, ["Secteurs", "Secteur"]),
        "responsibility": find_column(
            header, ["Responsabilité", "Responsabilite"], required=False
        ),
        "action": find_column(
            header, ["Action Déploiement", "Action Deploiement"], required=False
        ),
        "qualif": find_column(
            header, ["Qualif", "Qualification"], required=False
        ),
        "degraded_bands": find_column(
            header,
            [
                "Bandes débit Dégradé",
                "Bandes débit Dégradées",
                "Bandes debit Degrade",
            ],
            required=False,
        ),
        "charged_bands": find_column(
            header,
            ["Bandes Chargées", "Bandes Chargees"],
            required=False,
        ),
        "date_cols": [],
        "bands": find_band_groups(top_rows, header_idx),
    }

    for name in (
        "Prod Optim",
        "Prod Maintenance",
        "Prod Déploiement",
        "Prod Deploiement",
        "Prod Ingénierie",
        "Prod Ingenierie",
    ):
        idx = find_column(header, [name], required=False)
        if idx is not None and idx not in structure["date_cols"]:
            structure["date_cols"].append(idx)

    return structure


def extract_data(path):
    """
    Lecture read_only :
    - ne charge pas les centaines de milliers de lignes vides en mémoire ;
    - s'arrête après MAX_CONSECUTIVE_EMPTY_ROWS lignes sans secteur.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = find_analysis_sheet(wb)

    top_rows = read_top_rows(ws, 20)
    structure = detect_structure(top_rows)

    data = {}
    duplicates = Counter()
    empty_run = 0

    start_excel_row = structure["header_excel_row"] + 1

    for excel_row, row in enumerate(
        ws.iter_rows(min_row=start_excel_row, values_only=True),
        start=start_excel_row,
    ):
        sector = normalize_sector(safe_get(row, structure["sector"]))

        if not sector:
            empty_run += 1
            if empty_run >= MAX_CONSECUTIVE_EMPTY_ROWS:
                break
            continue

        empty_run = 0
        duplicates[sector] += 1

        if sector in data:
            continue

        record = {
            "excel_row": excel_row,
            "sector": safe_get(row, structure["sector"]),
            "responsibility": safe_get(row, structure["responsibility"]),
            "action": safe_get(row, structure["action"]),
            "qualif": safe_get(row, structure["qualif"]),
            "degraded_bands": safe_get(row, structure["degraded_bands"]),
            "charged_bands": safe_get(row, structure["charged_bands"]),
            "dates": [
                safe_get(row, c)
                for c in structure["date_cols"]
                if c is not None
            ],
            "bands": {},
        }

        for band in BANDS:
            record["bands"][band] = {
                "dl": safe_get(row, structure["bands"][band]["dl"]),
                "prb": safe_get(row, structure["bands"][band]["prb"]),
            }

        data[sector] = record

    wb.close()

    duplicate_keys = [k for k, n in duplicates.items() if n > 1]
    return data, duplicate_keys


def latest_action_date(record):
    dates = []
    for v in record.get("dates", []):
        if isinstance(v, datetime):
            dates.append(v)
        elif hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
            try:
                dates.append(datetime(v.year, v.month, v.day))
            except Exception:
                pass

    return max(dates) if dates else None


# ============================================================
# CALCUL TRD / IEA
# ============================================================

def calculate_sector(before, after):
    initial_degraded = clean_text(before.get("degraded_bands"))
    after_degraded = clean_text(after.get("degraded_bands"))
    initial_charged = clean_text(before.get("charged_bands"))

    dl_targets = band_set(initial_degraded)
    prb_targets = band_set(initial_charged)

    band_results = {}

    gap_dl_before = 0.0
    gap_dl_after = 0.0
    gap_prb_before = 0.0
    gap_prb_after = 0.0

    dl_measurable = False
    prb_measurable = False

    for band in BANDS:
        bdl = to_float(before["bands"][band]["dl"])
        adl = to_float(after["bands"][band]["dl"])
        bprb = to_float(before["bands"][band]["prb"])
        aprb = to_float(after["bands"][band]["prb"])

        band_results[band] = {
            "dl_before": bdl,
            "dl_after": adl,
            "gain_dl": safe_gain(bdl, adl),
            "prb_before": bprb,
            "prb_after": aprb,
            "relief_prb": prb_relief(bprb, aprb),
        }

        # Débit : gap = seuil - débit, uniquement sur la dégradation initiale.
        if band in dl_targets and bdl is not None:
            initial_gap = max(0.0, DL_THRESHOLDS[band] - bdl)

            if initial_gap > 0:
                dl_measurable = True
                gap_dl_before += initial_gap
                gap_dl_after += (
                    max(0.0, DL_THRESHOLDS[band] - adl)
                    if adl is not None
                    else DL_THRESHOLDS[band]
                )

        # PRB : gap = PRB - seuil, uniquement si la bande était chargée.
        if band in prb_targets and bprb is not None:
            initial_gap = max(0.0, bprb - PRB_THRESHOLD)

            if initial_gap > 0:
                prb_measurable = True
                gap_prb_before += initial_gap
                gap_prb_after += (
                    max(0.0, aprb - PRB_THRESHOLD)
                    if aprb is not None
                    else max(0.0, 100.0 - PRB_THRESHOLD)
                )

    trd_dl = (
        1.0 - gap_dl_after / gap_dl_before
        if dl_measurable and gap_dl_before > 0
        else None
    )

    trd_prb = (
        1.0 - gap_prb_after / gap_prb_before
        if prb_measurable and gap_prb_before > 0
        else None
    )

    if trd_dl is not None and trd_prb is not None:
        iea = trd_dl * WEIGHT_TRD_DL + trd_prb * WEIGHT_TRD_PRB
    elif trd_dl is not None:
        iea = trd_dl
    elif trd_prb is not None:
        iea = trd_prb
    else:
        iea = None

    action_status = classify_iea(iea)
    degradation_status = clean_text(after.get("qualif"))
    conclusion = build_conclusion(
        degradation_status, action_status, iea
    )

    if initial_degraded and after_degraded:
        if initial_degraded == after_degraded:
            evolution = "Même bande(s)"
        else:
            evolution = (
                f"Déplacement: {initial_degraded} → {after_degraded}"
            )
    else:
        evolution = ""

    return {
        "initial_degraded": initial_degraded,
        "after_degraded": after_degraded,
        "initial_charged": initial_charged,
        "bands": band_results,
        "gap_dl_before": gap_dl_before if dl_measurable else None,
        "gap_dl_after": gap_dl_after if dl_measurable else None,
        "trd_dl": trd_dl,
        "gap_prb_before": gap_prb_before if prb_measurable else None,
        "gap_prb_after": gap_prb_after if prb_measurable else None,
        "trd_prb": trd_prb,
        "iea": iea,
        "action_status": action_status,
        "degradation_status": degradation_status,
        "conclusion": conclusion,
        "evolution": evolution,
    }


# ============================================================
# EXCEL RESULTAT
# ============================================================

def result_headers():
    headers = [
        "Secteur",
        "Responsabilité",
        "Action",
        "Statut Dégradation",
        "Date Action",
        "Dégradation initiale",
        "Dégradation après",
        "Bandes chargées initiales",
    ]

    for band in BANDS:
        headers.extend([
            f"{band} DL Avant",
            f"{band} DL Après",
            f"Gain DL {band} %",
            f"{band} PRB Avant",
            f"{band} PRB Après",
            f"Soulagement PRB {band} pts",
        ])

    headers.extend([
        "Gap DL Avant",
        "Gap DL Après",
        "TRD Débit",
        "Gap PRB Avant",
        "Gap PRB Après",
        "TRD PRB",
        "IEA Action",
        "Statut Action",
        "Conclusion",
        "Confiance mesure",
        "Evolution bande dégradée",
    ])

    return headers


def style_result_sheet(ws):
    blue = "2F5597"
    white = "FFFFFF"

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 35

    header_idx = {
        clean_text(cell.value): cell.column
        for cell in ws[1]
    }

    for col in range(1, ws.max_column + 1):
        header = clean_text(ws.cell(1, col).value)
        width = 14

        if header == "Secteur":
            width = 42
        elif header in (
            "Responsabilité",
            "Statut Action",
            "Statut Dégradation",
        ):
            width = 25
        elif header in (
            "Action",
            "Conclusion",
            "Confiance mesure",
            "Evolution bande dégradée",
        ):
            width = 38
        elif "Dégradation" in header or "Bandes chargées" in header:
            width = 25

        ws.column_dimensions[get_column_letter(col)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = clean_text(ws.cell(1, cell.column).value)

            if (
                "Gain DL" in header
                or header.startswith("TRD ")
                or header == "IEA Action"
            ):
                cell.number_format = "0.0%"

            elif (
                "PRB" in header
                or "DL Avant" in header
                or "DL Après" in header
                or header.startswith("Gap ")
            ):
                cell.number_format = "0.00"

            elif header == "Date Action":
                cell.number_format = "dd/mm/yyyy"

    iea_col = header_idx.get("IEA Action")

    if iea_col and ws.max_row >= 2:
        rng = (
            f"{get_column_letter(iea_col)}2:"
            f"{get_column_letter(iea_col)}{ws.max_row}"
        )
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="num",
                start_value=-0.20,
                start_color="F4CCCC",
                mid_type="num",
                mid_value=0.20,
                mid_color="FFF2CC",
                end_type="num",
                end_value=0.80,
                end_color="D9EAD3",
            ),
        )


def create_dashboard(wb, result_ws):
    ws = wb.create_sheet("Dashboard_TRD")

    dark_blue = "17365D"
    blue = "2F5597"
    light_blue = "D9EAF7"
    white = "FFFFFF"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Dashboard – Efficacité des actions correctives"
    ws["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    ws["A1"].font = Font(color=white, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    headers = [clean_text(c.value) for c in result_ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    status_counts = Counter()
    degradation_counts = Counter()

    resp_total = Counter()
    resp_iea = defaultdict(list)
    resp_normalized = Counter()
    resp_persist_improved = Counter()
    resp_regression = Counter()

    all_iea = []

    for r in range(2, result_ws.max_row + 1):
        status = clean_text(
            result_ws.cell(r, idx["Statut Action"]).value
        )
        degradation = clean_text(
            result_ws.cell(r, idx["Statut Dégradation"]).value
        )
        responsibility = (
            clean_text(result_ws.cell(r, idx["Responsabilité"]).value)
            or "Non renseigné"
        )
        iea = to_float(
            result_ws.cell(r, idx["IEA Action"]).value
        )

        status_counts[status] += 1
        degradation_counts[degradation] += 1
        resp_total[responsibility] += 1

        if iea is not None:
            all_iea.append(iea)
            resp_iea[responsibility].append(iea)

        if degradation.lower() == "normalisé":
            resp_normalized[responsibility] += 1

        if (
            degradation.lower() == "persiste"
            and iea is not None
            and iea >= IEA_PARTIAL
        ):
            resp_persist_improved[responsibility] += 1

        if status == "Régression":
            resp_regression[responsibility] += 1

    total = result_ws.max_row - 1
    normalises = degradation_counts.get("Normalisé", 0)
    persistants_ameliores = sum(resp_persist_improved.values())
    regressions = status_counts.get("Régression", 0)
    avg_iea = sum(all_iea) / len(all_iea) if all_iea else None

    cards = [
        ("A3", "Secteurs analysés", total),
        ("C3", "Normalisés", normalises),
        ("E3", "Persistants améliorés", persistants_ameliores),
        ("G3", "Régressions", regressions),
    ]

    for cell_ref, label, value in cards:
        label_cell = ws[cell_ref]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=light_blue)
        label_cell.font = Font(bold=True, color=dark_blue)
        label_cell.alignment = Alignment(horizontal="center")

        value_cell = ws.cell(
            label_cell.row + 1,
            label_cell.column,
        )
        value_cell.value = value
        value_cell.font = Font(bold=True, size=14)
        value_cell.alignment = Alignment(horizontal="center")

    ws["A6"] = "IEA moyen"
    ws["A6"].fill = PatternFill("solid", fgColor=light_blue)
    ws["A6"].font = Font(bold=True, color=dark_blue)
    ws["B6"] = avg_iea
    ws["B6"].number_format = "0.0%"
    ws["B6"].font = Font(bold=True, size=14)

    # Répartition des statuts.
    ws["A9"] = "Statut Action"
    ws["B9"] = "Nombre"

    for ref in ("A9", "B9"):
        ws[ref].fill = PatternFill("solid", fgColor=blue)
        ws[ref].font = Font(color=white, bold=True)

    status_order = [
        "Très efficace",
        "Efficace",
        "Amélioration partielle",
        "Impact faible / non significatif",
        "Régression",
        "Non mesurable",
    ]

    for i, status in enumerate(status_order, start=10):
        ws.cell(i, 1).value = status
        ws.cell(i, 2).value = status_counts.get(status, 0)

    # Responsabilités.
    start = 18
    resp_headers = [
        "Responsabilité",
        "Total",
        "Mesurables",
        "IEA moyen",
        "Normalisés",
        "Persistants améliorés",
        "Régressions",
    ]

    for c, header in enumerate(resp_headers, start=1):
        cell = ws.cell(start, c)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for rr, resp in enumerate(
        sorted(resp_total),
        start=start + 1,
    ):
        vals = resp_iea.get(resp, [])

        ws.cell(rr, 1).value = resp
        ws.cell(rr, 2).value = resp_total[resp]
        ws.cell(rr, 3).value = len(vals)
        ws.cell(rr, 4).value = (
            sum(vals) / len(vals)
            if vals else None
        )
        ws.cell(rr, 4).number_format = "0.0%"
        ws.cell(rr, 5).value = resp_normalized[resp]
        ws.cell(rr, 6).value = resp_persist_improved[resp]
        ws.cell(rr, 7).value = resp_regression[resp]

    # Graphique.
    chart = BarChart()
    chart.type = "col"
    chart.title = "Répartition des statuts d'action"
    chart.y_axis.title = "Nombre de secteurs"
    chart.style = 10
    chart.height = 8
    chart.width = 15

    data = Reference(
        ws,
        min_col=2,
        min_row=9,
        max_row=9 + len(status_order),
    )
    cats = Reference(
        ws,
        min_col=1,
        min_row=10,
        max_row=9 + len(status_order),
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D9")

    for col, width in {
        "A": 32,
        "B": 14,
        "C": 16,
        "D": 14,
        "E": 18,
        "F": 22,
        "G": 14,
        "H": 14,
    }.items():
        ws.column_dimensions[col].width = width


def create_parameters_sheet(wb):
    ws = wb.create_sheet("Parametres_TRD")

    rows = [
        ["Paramètre", "Valeur", "Commentaire"],
        ["Seuil débit L1800 (Mbps)", DL_THRESHOLDS["L1800"], "Modifiable dans le script"],
        ["Seuil débit L2100 (Mbps)", DL_THRESHOLDS["L2100"], "Modifiable dans le script"],
        ["Seuil débit L2600 (Mbps)", DL_THRESHOLDS["L2600"], "Modifiable dans le script"],
        ["Seuil débit L800 (Mbps)", DL_THRESHOLDS["L800"], "Modifiable dans le script"],
        ["Seuil PRB DL BH (%)", PRB_THRESHOLD, "Gap PRB comptabilisé au-dessus de ce seuil"],
        ["Poids TRD Débit", WEIGHT_TRD_DL, "IEA si Débit et PRB sont mesurables"],
        ["Poids TRD PRB", WEIGHT_TRD_PRB, "IEA si Débit et PRB sont mesurables"],
        ["IEA Très efficace", IEA_VERY_EFFECTIVE, ""],
        ["IEA Efficace", IEA_EFFECTIVE, ""],
        ["IEA Amélioration partielle", IEA_PARTIAL, ""],
        ["IEA Régression", IEA_REGRESSION, ""],
        [],
        [
            "Méthode",
            "",
            "TRD = part du gap initial résorbée. "
            "Un secteur peut rester Persiste avec un TRD/IEA positif.",
        ],
        [
            "Confiance",
            "",
            "Avec deux snapshots, la mesure reste indicative. "
            "Pour la rendre robuste : mêmes BH/heures + Traffic DL "
            "et utilisateurs comparables.",
        ],
    ]

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.font = Font(color="FFFFFF", bold=True)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 90

    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )


def generate_result(before_path, after_path, output_path):
    print("\nLecture du fichier AVANT...")
    before_data, duplicate_before = extract_data(before_path)

    print("Lecture du fichier APRES...")
    after_data, duplicate_after = extract_data(after_path)

    common = sorted(set(before_data) & set(after_data))

    if not common:
        raise ValueError(
            "Aucun secteur commun trouvé entre AVANT et APRES."
        )

    print(f"Secteurs communs : {len(common)}")
    print("Calcul TRD / IEA...")

    wb = Workbook()
    result_ws = wb.active
    result_ws.title = "Comparatif_TRD"
    result_ws.append(result_headers())

    for n, sector_key in enumerate(common, start=1):
        before = before_data[sector_key]
        after = after_data[sector_key]

        m = calculate_sector(before, after)

        row = [
            after.get("sector"),
            after.get("responsibility"),
            after.get("action"),
            m["degradation_status"],
            latest_action_date(after),
            m["initial_degraded"],
            m["after_degraded"],
            m["initial_charged"],
        ]

        for band in BANDS:
            b = m["bands"][band]
            row.extend([
                b["dl_before"],
                b["dl_after"],
                b["gain_dl"],
                b["prb_before"],
                b["prb_after"],
                b["relief_prb"],
            ])

        confidence = (
            "Moyenne – 2 snapshots, trafic non contrôlé"
            if m["iea"] is not None
            else "Faible / non mesurable"
        )

        row.extend([
            m["gap_dl_before"],
            m["gap_dl_after"],
            m["trd_dl"],
            m["gap_prb_before"],
            m["gap_prb_after"],
            m["trd_prb"],
            m["iea"],
            m["action_status"],
            m["conclusion"],
            confidence,
            m["evolution"],
        ])

        result_ws.append(row)

        if n % 500 == 0:
            print(f"  {n}/{len(common)} secteurs...")

    style_result_sheet(result_ws)
    create_dashboard(wb, result_ws)
    create_parameters_sheet(wb)

    wb.save(output_path)

    print("\nRESULTAT GENERE")
    print(f"Fichier : {output_path}")
    print(f"Secteurs comparés : {len(common)}")

    if duplicate_before:
        print(
            f"Attention : {len(duplicate_before)} secteur(s) "
            "en doublon dans AVANT."
        )

    if duplicate_after:
        print(
            f"Attention : {len(duplicate_after)} secteur(s) "
            "en doublon dans APRES."
        )

    return output_path


# ============================================================
# FENETRES DE SELECTION
# ============================================================

def choose_files_gui():
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
    except ImportError:
        raise RuntimeError(
            "Tkinter n'est pas disponible. "
            "Utiliser --before / --after / --output."
        )

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    messagebox.showinfo(
        "Comparateur TRD",
        "Sélectionner le fichier AVANT.",
    )

    before = filedialog.askopenfilename(
        title="1/2 - Fichier AVANT",
        filetypes=[
            ("Fichiers Excel", "*.xlsx *.xlsm"),
            ("Tous les fichiers", "*.*"),
        ],
    )

    if not before:
        root.destroy()
        return None, None, None

    messagebox.showinfo(
        "Comparateur TRD",
        "Sélectionner maintenant le fichier APRES.",
    )

    after = filedialog.askopenfilename(
        title="2/2 - Fichier APRES",
        filetypes=[
            ("Fichiers Excel", "*.xlsx *.xlsm"),
            ("Tous les fichiers", "*.*"),
        ],
    )

    if not after:
        root.destroy()
        return None, None, None

    default_name = (
        Path(after).stem
        + "_Resultat_TRD_"
        + datetime.now().strftime("%Y%m%d_%H%M")
        + ".xlsx"
    )

    output = filedialog.asksaveasfilename(
        title="Enregistrer le résultat",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Fichier Excel", "*.xlsx")],
    )

    root.destroy()

    if not output:
        return None, None, None

    return before, after, output


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Compare deux consolidations Avant/Après "
            "et calcule TRD / IEA."
        )
    )

    parser.add_argument("--before", help="Fichier AVANT")
    parser.add_argument("--after", help="Fichier APRES")
    parser.add_argument("--output", help="Fichier résultat")

    args = parser.parse_args()

    if args.before and args.after:
        before = args.before
        after = args.after

        output = (
            args.output
            if args.output
            else str(
                Path(after).with_name(
                    Path(after).stem
                    + "_Resultat_TRD.xlsx"
                )
            )
        )
    else:
        before, after, output = choose_files_gui()

        if not before or not after or not output:
            print("Opération annulée.")
            return

    try:
        generate_result(before, after, output)

    except Exception as exc:
        print("\nERREUR :", exc)

        try:
            import tkinter as tk
            from tkinter import messagebox

            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "Erreur Comparateur TRD",
                str(exc),
            )
            root.destroy()
        except Exception:
            pass

        raise


if __name__ == "__main__":
    main()
